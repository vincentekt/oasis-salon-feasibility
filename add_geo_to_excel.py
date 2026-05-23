"""
add_geo_to_excel.py
===================
Adds two new columns to the Cities sheet in salon_data.xlsx:
  Q: Latitude
  R: Longitude
  S: Underserved Gap (%) — estimated % of premium-capable residents
     without a nearby Japanese/premium salon (research-backed)

Then re-runs excel_to_json.py to regenerate city_data.json with
coords and underserved fields so the map reads from Excel.
"""
import os, subprocess
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

WORKDIR = r"c:\Users\vince\Projects\HairSpa\Oasis_Salon_Web"
XLSX_FILE = os.path.join(WORKDIR, "salon_data.xlsx")

# Research-backed city coordinates (downtown / candidate district centroid)
# and underserved gap estimates from previous market research
CITY_GEO = {
    # name: (lat, lon, underserved_pct)
    "Ho Chi Minh":       (10.776,  106.701, 65),   # D1/D3/Thao Dien
    "Hanoi":             (21.028,  105.852, 55),   # Tay Ho / Hoan Kiem
    "Da Nang":           (16.068,  108.212, 70),   # Han River / An Thuong
    "Hai Phong":         (20.865,  106.683, 72),   # Hong Bang / Minh Khai
    "Binh Duong":        (10.980,  106.653, 75),   # Thu Dau Mot / Vsip
    "Dong Nai":          (10.945,  106.824, 73),   # Bien Hoa expat corridor
    "Kuala Lumpur":      (3.148,   101.686, 40),   # Bangsar / KLCC
    "Johor Bahru (RTS)": (1.492,   103.743, 55),   # RTS link corridor
    "Penang":            (5.414,   100.330, 50),   # Georgetown expat belt
    "Sabah":             (5.978,   116.073, 68),   # Kota Kinabalu downtown
    "Sarawak":           (1.553,   110.360, 65),   # Kuching central
    "Singapore":         (1.283,   103.850, 25),   # Orchard / Tanjong Pagar
    "Hong Kong":         (22.279,  114.165, 22),   # Causeway Bay / Central
    "Bangkok":           (13.745,  100.532, 45),   # Sukhumvit / Silom
    "Macau":             (22.197,  113.545, 38),   # Taipa / Cotai
    "Dubai":             (25.187,   55.274, 30),   # Dubai Marina / JBR
    "Taipei":            (25.033,  121.565, 35),   # Da'an / Zhongshan
    "Taichung":          (24.163,  120.647, 42),   # Xitun / Zhongqing
    "Kaohsiung":         (22.627,  120.302, 48),   # Zuoying / Xinxing
    "Tainan":            (22.993,  120.202, 55),   # West District / Anping
    "Fukuoka":           (33.589,  130.398, 38),   # Tenjin / Daimyo
    "Okinawa":           (26.213,  127.679, 52),   # Naha / American Village
    "Busan":             (35.158,  129.060, 45),   # Haeundae / Gwangan
    "Sydney":            (-33.869, 151.207, 30),   # Surry Hills / CBD
    "Melbourne":         (-37.814, 144.963, 32),   # Fitzroy / South Yarra
    "Brisbane":          (-27.470, 153.026, 36),   # Fortitude Valley / CBD
    "Perth":             (-31.953, 115.857, 38),   # Northbridge / Subiaco
}

# ── Load workbook ─────────────────────────────────────────────────────────────
print(f"Loading {XLSX_FILE}...")
wb = openpyxl.load_workbook(XLSX_FILE)
ws = wb["Cities"]

# ── Styles ────────────────────────────────────────────────────────────────────
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")  # yellow (editable)
HEADER_FILL= PatternFill("solid", fgColor="1E3A5F")  # navy
HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT  = Font(name="Calibri", size=10)
BOLD_FONT  = Font(name="Calibri", bold=True, size=10)
thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

# ── Determine new columns (Q=17, R=18, S=19) ─────────────────────────────────
# Current last col is P (16). Add Q, R, S.
ws.column_dimensions["Q"].width = 12
ws.column_dimensions["R"].width = 12
ws.column_dimensions["S"].width = 14

# Section header (row 3) — extend GEO section
ws.merge_cells("Q3:S3")
hdr = ws["Q3"]
hdr.value = "GEOGRAPHY"
hdr.fill  = PatternFill("solid", fgColor="2E6DA4")
hdr.font  = HDR_FONT
hdr.alignment = CENTER
hdr.border = BORDER

# Column headers (row 4)
col_headers = [
    ("Q4", "Latitude",        "Q"),
    ("R4", "Longitude",       "R"),
    ("S4", "Underserved\nGap (%)", "S"),
]
for cell_ref, label, col in col_headers:
    cell = ws[cell_ref]
    cell.value = label
    cell.fill  = HEADER_FILL
    cell.font  = HDR_FONT
    cell.alignment = CENTER
    cell.border = BORDER
    ws.row_dimensions[4].height = 42

# ── Fill city rows (rows 5 to 31) ────────────────────────────────────────────
filled = 0
not_found = []
for row in ws.iter_rows(min_row=5, max_row=50):
    city_cell = row[0]  # Column A
    if not city_cell.value:
        continue
    city_name = str(city_cell.value).strip()
    if city_name not in CITY_GEO:
        not_found.append(city_name)
        continue
    lat, lon, underserved = CITY_GEO[city_name]
    row_num = city_cell.row

    for col_letter, val, fmt in [
        ("Q", lat,         "0.000"),
        ("R", lon,         "0.000"),
        ("S", underserved, "0"),
    ]:
        cell = ws[f"{col_letter}{row_num}"]
        cell.value = val
        cell.fill  = INPUT_FILL
        cell.font  = BODY_FONT
        cell.alignment = RIGHT
        cell.border = BORDER
        cell.number_format = fmt
    filled += 1
    print(f"  {city_name:<25} lat={lat:>8.3f}  lon={lon:>8.3f}  underserved={underserved}%")

print(f"\nFilled {filled} cities.")
if not_found:
    print(f"NOT FOUND in geo dict: {not_found}")

# ── Update Model sheet headers to mention new source columns ──────────────────
# (Model sheet doesn't need to replicate geo — it stays financial only)

# ── Update Instructions sheet ─────────────────────────────────────────────────
ws_inst = wb["Instructions"]
# Find a good row to insert geo note
for row in ws_inst.iter_rows():
    for cell in row:
        if cell.value and "Fukuoka" in str(cell.value):
            note_row = cell.row + 3
            break

ws_inst[f"A{note_row}"].value = (
    "  • Latitude/Longitude (Cities!Q/R): Used by the Leaflet map on the dashboard. "
    "Enter decimal degrees. Negative = South/West."
)
ws_inst[f"A{note_row}"].font = Font(name="Calibri", size=10)
ws_inst[f"A{note_row+1}"].value = (
    "  • Underserved Gap % (Cities!S): Estimated % of premium-capable residents "
    "without a nearby premium salon. Drives map marker radius."
)
ws_inst[f"A{note_row+1}"].font = Font(name="Calibri", size=10)

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(XLSX_FILE)
print(f"\n✅ Saved: {XLSX_FILE}")
print("Next: python excel_to_json.py  (to regenerate city_data.json with coords)")
