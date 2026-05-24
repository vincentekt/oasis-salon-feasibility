import openpyxl
import os

web_dir = r"c:\Users\vince\Projects\HairSpa\Oasis_Salon_Web"
path = os.path.join(web_dir, "salon_data.xlsx")

print(f"Opening workbook: {path}")
wb = openpyxl.load_workbook(path)
ws = wb["Cities"]

# Sabah is at row 14, column 19 (column S)
old_val = ws.cell(row=14, column=19).value
print(f"Old Sabah gap value: {old_val}")
ws.cell(row=14, column=19).value = 75
print(f"New Sabah gap value set to: {ws.cell(row=14, column=19).value}")

wb.save(path)
print("Workbook saved successfully.")
