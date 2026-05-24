import openpyxl
import os

xlsx_path = r"c:\Users\vince\Projects\HairSpa\Oasis_Salon_Web\salon_data.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb["Cities"]

# Print header
headers = [cell.value for cell in ws[4]]
print("Headers:", headers)

for row in ws.iter_rows(min_row=5, max_row=50, values_only=True):
    if row[0] and "Hong Kong" in str(row[0]):
        print("HK Data Row:")
        for h, val in zip(headers, row):
            print(f"  {h}: {val}")
