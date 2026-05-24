import openpyxl

xlsx_path = r"c:\Users\vince\Projects\HairSpa\Oasis_Salon_Web\salon_data.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
ws = wb["Cities"]

for r in range(1, 100):
    val = ws.cell(row=r, column=1).value
    if val == "Hong Kong":
        # Find column for Underserved Gap %
        for c in range(1, 30):
            header_val = ws.cell(row=4, column=c).value
            if header_val and "Underserved" in str(header_val):
                print(f"Row {r}, Column {c} ({openpyxl.utils.get_column_letter(c)}{r}): header='{header_val}', value='{ws.cell(row=r, column=c).value}'")
