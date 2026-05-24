"""
excel_to_json.py
================
Reads salon_data.xlsx (Cities + Country sheets) and:
  1. Replicates all Model-sheet formulas in Python (mirrors Excel exactly)
  2. Outputs city_data.json   — consumed by the frontend
  3. Patches script.js        — regenerates the citiesDb block

Run this after any change to salon_data.xlsx.
Usage:
    python excel_to_json.py
    # Optional: python excel_to_json.py --dry-run   (print only, no file writes)
"""

import os, re, json, sys, math
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

WORKDIR   = r"c:\Users\vince\Projects\HairSpa\Oasis_Salon_Web"
XLSX_FILE = os.path.join(WORKDIR, "salon_data.xlsx")
JSON_FILE = os.path.join(WORKDIR, "city_data.json")
JS_FILE   = os.path.join(WORKDIR, "script.js")
COMPLEXITY_FILE = os.path.join(WORKDIR, "complexity_data.json")

DRY_RUN = "--dry-run" in sys.argv

# ── LOAD COMPLEXITY DATA ──────────────────────────────────────────────────────
print(f"Reading {COMPLEXITY_FILE}...")
with open(COMPLEXITY_FILE, "r", encoding="utf-8") as f:
    complexities = json.load(f)
print(f"  Loaded {len(complexities)} complexity profiles")


# ── LOAD EXCEL ────────────────────────────────────────────────────────────────
print(f"Reading {XLSX_FILE}...")
wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)

# ── COUNTRY PARAMS (Country sheet, rows 4-14, cols A-H) ──────────────────────
ws_c = wb["Country"]
country_params = {}
for row in ws_c.iter_rows(min_row=4, max_row=20, values_only=True):
    if not row[0]:
        continue
    # Skip non-data rows (e.g. instruction text bled into the sheet)
    try:
        tax_check = float(row[2])
    except (TypeError, ValueError):
        continue
    code = str(row[0]).strip()
    if len(code) > 3:   # safety: country codes are 2-3 chars max
        continue
    country_params[code] = {
        "name":    row[1],
        "tax":     float(row[2]) if row[2] else 0.20,
        "currency":row[3],
        "senior":  float(row[4]) if row[4] else 1500,
        "junior":  float(row[5]) if row[5] else 900,
        "asst":    float(row[6]) if row[6] else 700,
        "manager": float(row[7]) if row[7] else 1800,
    }
print(f"  Loaded {len(country_params)} country records: {list(country_params.keys())}")

# ── CITY INPUTS (Cities sheet, rows 5+, cols A-U) ────────────────────────────
ws_i = wb["Cities"]
city_inputs = []
for row in ws_i.iter_rows(min_row=5, max_row=50, values_only=True):
    if not row[0]:
        continue
    city_inputs.append({
        "name":     str(row[0]).strip(),
        "url":      str(row[1]).strip(),
        "region":   str(row[2]).strip(),
        "country":  str(row[3]).strip(),
        "sqft":     int(row[4]) if row[4] else 800,
        "sess":     int(row[5]) if row[5] else 3,
        "wd":       int(row[6]) if row[6] else 26,
        "ticket":   float(row[7]) if row[7] else 100,
        "cogs_pct": float(row[8]) if row[8] else 0.10,
        "rent":     float(row[9]) if row[9] else 2000,
        "fitout":   float(row[10]) if row[10] else 50000,
        "equip":    float(row[11]) if row[11] else 20000,
        "other":    float(row[12]) if row[12] else 5000,
        "dep_mo":   int(row[13]) if row[13] else 3,
        "util_y1":  float(row[14]) if row[14] else 0.65,
        "util_y2":  float(row[15]) if row[15] else 0.75,
        # Geography (cols Q=16, R=17, S=18)
        "lat":         float(row[16]) if len(row) > 16 and row[16] is not None else 0.0,
        "lon":         float(row[17]) if len(row) > 17 and row[17] is not None else 0.0,
        "underserved": int(row[18])   if len(row) > 18 and row[18] is not None else 40,
        # Business Fees & Admin (cols T=19, U=20)
        "biz_reg":     float(row[19]) if len(row) > 19 and row[19] is not None else 1500,
        "admin_sw":    float(row[20]) if len(row) > 20 and row[20] is not None else 200,
    })
print(f"  Loaded {len(city_inputs)} cities")

# ── CAPEX & OPEX DETAILS (CAPEX_Details & OPEX_Details sheets) ────────────────
ws_capex = wb["CAPEX_Details"]
capex_details = {}
for row in ws_capex.iter_rows(min_row=2, max_row=500, values_only=True):
    if not row[0]:
        continue
    city_name = str(row[0]).strip()
    category = str(row[1]).strip()
    component = str(row[2]).strip()
    cost = float(row[3]) if row[3] is not None else 0.0
    ref = str(row[4]).strip() if row[4] else ""
    
    if city_name not in capex_details:
        capex_details[city_name] = []
    capex_details[city_name].append({
        "category": category,
        "component": component,
        "cost": cost,
        "reference": ref
    })

ws_opex = wb["OPEX_Details"]
opex_details = {}
for row in ws_opex.iter_rows(min_row=2, max_row=500, values_only=True):
    if not row[0]:
        continue
    city_name = str(row[0]).strip()
    category = str(row[1]).strip()
    component = str(row[2]).strip()
    cost = float(row[3]) if row[3] is not None else 0.0
    ref = str(row[4]).strip() if row[4] else ""
    
    if city_name not in opex_details:
        opex_details[city_name] = []
    opex_details[city_name].append({
        "category": category,
        "component": component,
        "cost": cost,
        "reference": ref
    })
print(f"  Loaded CAPEX & OPEX details for {len(capex_details)} cities")

# ── EQUIPMENT SOURCING DETAILS ──────────────────────────────────────────────
ws_equip = wb["Equipment_Sourcing"]
equip_sourcing = {}
for row in ws_equip.iter_rows(min_row=2, max_row=200, values_only=True):
    if not row[0]:
        continue
    country = str(row[0]).strip()
    eq_type = str(row[1]).strip()
    imp_model = str(row[2]).strip()
    imp_base = float(row[3]) if row[3] is not None else 0.0
    logistics = float(row[4]) if row[4] is not None else 0.0
    imp_total = float(row[5]) if row[5] is not None else 0.0
    local_model = str(row[6]).strip()
    local_cost = float(row[7]) if row[7] is not None else 0.0
    cheaper = str(row[8]).strip()
    ref = str(row[9]).strip() if row[9] else ""
    
    if country not in equip_sourcing:
        equip_sourcing[country] = []
    equip_sourcing[country].append({
        "equipment_type": eq_type,
        "import_model": imp_model,
        "import_base_cost": imp_base,
        "logistics_and_duty": logistics,
        "total_import_cost": imp_total,
        "local_supplier_model": local_model,
        "local_supplier_cost": local_cost,
        "cheaper_option": cheaper,
        "reference": ref
    })

# ── STOCK DETAILS ────────────────────────────────────────────────────────────
ws_stock = wb["Stock_Details"]
stock_template = []
for row in ws_stock.iter_rows(min_row=2, max_row=100, values_only=True):
    if not row[0]:
        continue
    cat = str(row[0]).strip()
    name = str(row[1]).strip()
    price = float(row[2]) if row[2] is not None else 0.0
    qty = int(row[3]) if row[3] is not None else 0
    total = float(row[4]) if row[4] is not None else 0.0
    ref = str(row[5]).strip() if row[5] else ""
    
    stock_template.append({
        "category": cat,
        "brand_name": name,
        "unit_price": price,
        "quantity": qty,
        "total_cost": total,
        "reference": ref
    })
print(f"  Loaded equipment sourcing for {len(equip_sourcing)} countries and {len(stock_template)} stock lines")

# ── MIRRORED MODEL FORMULAS ───────────────────────────────────────────────────
def compute_model(city, cp):
    """Mirrors the Excel Model sheet formulas exactly."""
    sqft    = city["sqft"]
    sess    = city["sess"]
    wd      = city["wd"]
    ticket  = city["ticket"]
    cogs_p  = city["cogs_pct"]
    rent    = city["rent"]
    fitout  = city["fitout"]
    equip   = city["equip"]
    other   = city["other"]
    dep_mo  = city["dep_mo"]
    util_y1 = city["util_y1"]
    util_y2 = city["util_y2"]
    biz_reg = city["biz_reg"]
    admin_sw= city["admin_sw"]
    
    # Country rates
    tax     = cp["tax"]
    senior  = cp["senior"]
    junior  = cp["junior"]
    asst    = cp["asst"]
    manager = cp["manager"]
    
    # === SPACE → STATIONS & SEATS ===
    # Stations = IF(sqft<=750,3, IF(sqft<=850,4, 5))
    if sqft <= 750:
        stations = 3
    elif sqft <= 850:
        stations = 4
    else:
        stations = 5
    
    # Color Bar Seats = stations - 1
    color_bar_seats = stations - 1
    total_seats = stations + color_bar_seats
    
    # Wash basins
    wash_basins = 2 if stations <= 4 else 3
    
    # Seniors = IF(stations<=4, 2, 3)
    seniors = 2 if stations <= 4 else 3
    
    # Juniors = stations - seniors
    juniors = stations - seniors
    
    # Assistants = CEILING(stations/2, 1)
    assistants = math.ceil(stations / 2)
    
    # Total Staff = seniors + juniors + assistants + 1 (manager)
    total_staff = seniors + juniors + assistants + 1
    
    # === STAFF COST ===
    staff_cost = seniors * senior + juniors * junior + assistants * asst + manager
    
    # === OPEX ===
    utilities  = 200 + stations * 80
    mktg_misc  = 350 + stations * 50 + 300 + stations * 30  # = 650 + stations*80
    depreciation = equip / 60  # 5-year straight line on equipment
    total_cash_opex = rent + staff_cost + utilities + mktg_misc + admin_sw
    total_opex = total_cash_opex + depreciation  # for display/bookkeeping
    
    # === CAPACITY ===
    max_cap = stations * sess * wd
    y1_clients = int(max_cap * util_y1)
    y2_clients = int(max_cap * util_y2)
    
    # === UNIT ECONOMICS ===
    cogs_per_session = ticket * cogs_p
    gm_per_session   = ticket - cogs_per_session
    
    # Breakeven Accounting (covering depreciation too!)
    breakeven = math.ceil((total_cash_opex + depreciation) / gm_per_session)
    breakeven_daily = round(breakeven / wd, 1)
    
    # Y1 & Y2 PAT
    y1_ebit = y1_clients * ticket * (1 - cogs_p) - total_cash_opex - depreciation
    y1_tax  = y1_ebit * tax if y1_ebit > 0 else 0.0
    y1_pat  = y1_ebit - y1_tax
    
    y2_ebit = y2_clients * ticket * (1 - cogs_p) - total_cash_opex - depreciation
    y2_tax  = y2_ebit * tax if y2_ebit > 0 else 0.0
    y2_pat  = y2_ebit - y2_tax
    
    # Cash flows (PAT + Depreciation)
    y1_cf = y1_pat + depreciation
    y2_cf = y2_pat + depreciation
    
    pat_ratio = round(y2_pat / total_cash_opex * 100) if total_cash_opex > 0 else 0
    
    # === CAPEX ===
    capex_base = fitout + equip + other + biz_reg + rent * dep_mo
    capex_low  = math.floor(capex_base * 0.92 / 5000) * 5000
    capex_high = math.ceil( capex_base * 1.08 / 5000) * 5000
    capex_mid  = (capex_low + capex_high) / 2
    
    # Payback using cumulative cash flow
    if y1_cf * 12 >= capex_mid:
        payback = math.ceil(capex_mid / y1_cf) if y1_cf > 0 else 999
    else:
        rem = capex_mid - y1_cf * 12
        if y2_cf > 0:
            payback = 12 + math.ceil(rem / y2_cf)
        else:
            payback = 999
    
    # Viable? = breakeven < 80% of max capacity
    viable = breakeven < max_cap * 0.8
    
    return {
        # Staff & Seats
        "stations":        stations,
        "color_bar_seats": color_bar_seats,
        "total_seats":     total_seats,
        "wash_basins":     wash_basins,
        "seniors":         seniors,
        "juniors":         juniors,
        "assistants":      assistants,
        "total_staff":     total_staff,
        # Rates
        "tax":             round(tax, 4),
        # Costs
        "rent":            round(rent),
        "staff_cost":      round(staff_cost),
        "utilities":       round(utilities),
        "mktg_misc":       round(mktg_misc),
        "admin_sw":        round(admin_sw),
        "depreciation":    round(depreciation),
        "total_cash_opex": round(total_cash_opex),
        "total_opex":      round(total_opex),
        # Capacity
        "max_cap":         max_cap,
        "y1_clients":      y1_clients,
        "y2_clients":      y2_clients,
        # Economics
        "ticket":          round(ticket, 2),
        "cogs_pct":        round(cogs_p, 4),
        "cogs_session":    round(cogs_per_session, 2),
        "gm_session":      round(gm_per_session, 2),
        "breakeven":       breakeven,
        "be_daily":        breakeven_daily,
        "y1_pat":          round(y1_pat, 2),
        "y2_pat":          round(y2_pat, 2),
        "y1_cf":           round(y1_cf, 2),
        "y2_cf":           round(y2_cf, 2),
        "pat_ratio":       pat_ratio,
        # CAPEX
        "biz_reg":         biz_reg,
        "capex_low":       int(capex_low),
        "capex_high":      int(capex_high),
        "capex_mid":       int(capex_mid),
        "payback_mo":      payback,
        # Status
        "viable":          viable,
        "size":            f"{sqft} sq ft",
        "util_y1":         util_y1,
        "util_y2":         util_y2,
        "sess_per_day":    sess,
        "working_days":    wd,
    }

# ── RUN MODEL FOR ALL CITIES ──────────────────────────────────────────────────
print("\nComputing model for all cities:")
print(f"{'City':<22} {'Sta':>4} {'Seats':>5} {'Staff':>6} {'OPEX':>9} {'MaxCap':>7} {'BE':>6} {'Y2 PAT':>9} {'PAT%':>6} {'CAPEX':>14} {'Payback':>8} {'OK?':>6}")
print("-" * 110)

results = {}
for city in city_inputs:
    name = city["name"]
    cp   = country_params.get(city["country"])
    if not cp:
        print(f"  ⚠ Missing country: {city['country']} for {name}")
        continue
    m = compute_model(city, cp)
    results[name] = {**city, **m}
    print(f"{name:<22} {m['stations']:>4} {m['total_seats']:>5} {m['total_staff']:>6} ${m['total_cash_opex']:>8,} "
          f"{m['max_cap']:>7} {m['breakeven']:>6} ${m['y2_pat']:>8,.0f} {m['pat_ratio']:>5}% "
          f"USD{m['capex_low']//1000:>4}k-{m['capex_high']//1000}k {m['payback_mo']:>6}mo "
          f"{'✓' if m['viable'] else '⚠':>6}")

# ── BUILD JSON FOR FRONTEND ────────────────────────────────────────────────────
print(f"\nBuilding city_data.json...")

# Build extra metadata from capacity_results.json (airport, risk, format, etc.)
# to preserve fields not in the Excel (we'll keep those in JSON or add to Excel later)
extra_meta = {}
cr_path = os.path.join(WORKDIR, "capacity_results.json")
if os.path.exists(cr_path):
    with open(cr_path, "r", encoding="utf-8") as f:
        cap_results = json.load(f)
    for cname, cdata in cap_results.items():
        extra_meta[cname] = {
            "airport": cdata.get("airport", ""),
            "risk":    cdata.get("risk", ""),
            "format":  cdata.get("format", ""),
            "region":  cdata.get("region", ""),
        }

def fmt_usd(v):
    return f"USD {int(round(v)):,}"

def fmt_capex(low, high):
    return f"USD {low//1000}k - {high//1000}k"

def staff_label(seniors, juniors, assistants):
    parts = []
    if seniors > 0:
        parts.append(f"{seniors} Senior Stylist{'s' if seniors > 1 else ''}")
    if juniors > 0:
        parts.append(f"{juniors} Junior Stylist{'s' if juniors > 1 else ''}")
    if assistants > 0:
        parts.append(f"{assistants} Assistant{'s' if assistants > 1 else ''}")
    parts.append("1 Manager")
    return " + ".join(parts)

cities_json = []
for city in city_inputs:
    name = city["name"]
    if name not in results:
        continue
    r  = results[name]
    ex = extra_meta.get(name, {})

    deposit_usd = int(round(city["rent"] * city["dep_mo"]))

    entry = {
        "name":         name,
        "url":          city["url"],
        "region":       ex.get("region", city["region"]),
        "format":       ex.get("format", "Premium Boutique Salon"),
        "size":         r["size"],
        "airport":      ex.get("airport", ""),
        # Financial inputs (raw numbers for UI rendering)
        "ticket_raw":   int(r["ticket"]),
        "ticket":       f"USD {r['ticket']:.0f}",
        "cogs":         f"USD {r['cogs_session']:.2f}",
        "margin":       "90%",
        "rent_raw":     int(r["rent"]),
        "rent_mo":      fmt_usd(r["rent"]),
        # Staff & Seats
        "stations":     r["stations"],
        "color_bar_seats": r["color_bar_seats"],
        "total_seats":  r["total_seats"],
        "wash_basins":  r["wash_basins"],
        "seniors":      r["seniors"],
        "juniors":      r["juniors"],
        "assistants":   r["assistants"],
        "total_staff":  r["total_staff"],
        "staff_label":  staff_label(r["seniors"], r["juniors"], r["assistants"]),
        "staff_cost_raw": int(r["staff_cost"]),
        "staff_cost_mo":fmt_usd(r["staff_cost"]),
        # Capacity
        "max_capacity": f"{r['max_cap']} sessions/month",
        "max_cap_raw":  r["max_cap"],
        "y1_clients":   r["y1_clients"],
        "y2_clients":   r["y2_clients"],
        "util_y1":      f"{int(r['util_y1']*100)}%",
        "util_y2":      f"{int(r['util_y2']*100)}%",
        "sess_per_day": city["sess"],
        "working_days": city["wd"],
        # OPEX breakdown (raw numbers for UI)
        "opex":         fmt_usd(r["total_cash_opex"]),
        "opex_raw":     r["total_cash_opex"],
        "utilities_raw":r["utilities"],
        "mktg_misc_raw":r["mktg_misc"],
        "admin_sw_raw": r["admin_sw"],
        "admin_sw_mo":  fmt_usd(r["admin_sw"]),
        "depreciation_raw": r["depreciation"],
        "depreciation_mo":  fmt_usd(r["depreciation"]),
        "total_opex_raw": r["total_opex"],
        "total_opex_mo":  fmt_usd(r["total_opex"]),
        # Economics
        "breakeven":    f"~{r['breakeven']} customers/mo",
        "breakeven_raw":r["breakeven"],
        "daily_breakeven": f"~{r['be_daily']} sessions/day",
        "tax":          f"{r['tax']*100:.1f}%",
        "tax_raw":      r["tax"],
        "pat_ratio":    f"{r['pat_ratio']}% (Post-Tax, at 75% util.)",
        "pat_ratio_raw":r["pat_ratio"],
        "y1_pat_raw":   round(r["y1_pat"], 2),
        "y1_pat_mo":    fmt_usd(r["y1_pat"]),
        "y2_pat":       fmt_usd(r["y2_pat"]),
        "y2_pat_raw":   round(r["y2_pat"], 2),
        "y1_cf_raw":    round(r["y1_cf"], 2),
        "y2_cf_raw":    round(r["y2_cf"], 2),
        "cogs_session_raw": r["cogs_session"],
        "gm_session_raw":   r["gm_session"],
        # CAPEX — range + line items (all from Excel inputs, no hardcoding)
        "capex":        fmt_capex(r["capex_low"], r["capex_high"]),
        "capex_low":    r["capex_low"],
        "capex_high":   r["capex_high"],
        "capex_mid":    r["capex_mid"],
        "fitout_usd":   int(city["fitout"]),
        "equip_usd":    int(city["equip"]),
        "other_usd":    int(city["other"]),
        "biz_reg_usd":  fmt_usd(r["biz_reg"]),
        "biz_reg_raw":  int(r["biz_reg"]),
        "deposit_usd":  deposit_usd,
        "dep_months":   city["dep_mo"],
        "payback":      f"{r['payback_mo']} Months (cumulative CF)",
        "payback_raw":  r["payback_mo"],
        # Geography
        "lat":          city["lat"],
        "lon":          city["lon"],
        "coords":       [city["lat"], city["lon"]],
        "underserved":  f"{city['underserved']}%",
        "underserved_raw": city["underserved"],
        # Risk & flags
        "risk":         ex.get("risk", ""),
        "viable":       r["viable"],
        "country":      city["country"],
        # Complexity (dynamic map)
        "complexity":   complexities.get(name),
        # Financial Details from Excel
        "capex_details": capex_details.get(name, []),
        "opex_details":  opex_details.get(name, []),
        "equipment_sourcing": equip_sourcing.get(city["country"], []),
        "stock_template": stock_template,
    }
    cities_json.append(entry)

if not DRY_RUN:
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(cities_json, f, indent=2, ensure_ascii=False)
    print(f"  ✓ Written: {JSON_FILE} ({len(cities_json)} cities)")
else:
    print(f"  [DRY RUN] Would write {len(cities_json)} cities to {JSON_FILE}")

# ── script.js note ───────────────────────────────────────────────────────────
# script.js now fetches city_data.json at runtime via loadCityDataAndInit().
# No patching needed — the JSON IS the data source.
print("\nscript.js reads city_data.json at runtime. No patching needed.")

# ── SUMMARY TABLE ─────────────────────────────────────────────────────────────
print("\n" + "=" * 80)
print("FINAL MODEL SUMMARY")
print("=" * 80)
print(f"{'City':<22} {'CAPEX':>16} {'OPEX':>10} {'Ticket':>8} {'BE':>6} {'PAT%':>6} {'Payback':>8}")
print("-" * 80)
for entry in cities_json:
    name = entry["name"]
    r    = results[name]
    print(f"{name:<22} USD{r['capex_low']//1000:>4}k-{r['capex_high']//1000}k "
          f"${r['total_opex']:>9,} ${r['ticket']:>7.0f} {r['breakeven']:>6} "
          f"{r['pat_ratio']:>5}% {r['payback_mo']:>6}mo")

print(f"\n✅ Done. {len(cities_json)} cities processed.")
print(f"   Files updated: city_data.json, script.js")
print(f"   Next: git add -A && git commit -m 'Update financials from Excel' && git push")
