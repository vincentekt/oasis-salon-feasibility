"""
create_salon_excel.py
=====================
Creates salon_data.xlsx — the single source of truth for all 27 city financials.

Architecture:
  - "Country" sheet:   Tax rates + salary data per country (editable)
  - "Cities" sheet:    All raw inputs per city (editable, yellow cells), including coords and biz fees.
  - "Model" sheet:     Calculated metrics using Excel formulas (blue, read-only), now with Depreciation,
                       Admin & Software OPEX, Business Registration CAPEX, and exact Cash Flow payback logic.
  - "Instructions" sheet: How to use + update workflow

After editing the Excel:
  Run: python excel_to_json.py
  This regenerates city_data.json.
  Then: git add -A && git commit && git push
"""

import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

WORKDIR = r"c:\Users\vince\Projects\HairSpa\Oasis_Salon_Web"
OUT_FILE = os.path.join(WORKDIR, "salon_data.xlsx") if "os" in globals() else "salon_data.xlsx"
# We'll use absolute path for safety
import os
OUT_FILE = os.path.join(WORKDIR, "salon_data.xlsx")

wb = openpyxl.Workbook()

# ── STYLES ────────────────────────────────────────────────────────────────────
INPUT_FILL    = PatternFill("solid", fgColor="FFF9C4")  # yellow
CALC_FILL     = PatternFill("solid", fgColor="DDEEFF")  # light blue
HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")  # dark navy
SUBHDR_FILL   = PatternFill("solid", fgColor="2E6DA4")  # medium blue
COUNTRY_FILL  = PatternFill("solid", fgColor="E8F5E9")  # light green
WARN_FILL     = PatternFill("solid", fgColor="FFCCCC")  # light red
OK_FILL       = PatternFill("solid", fgColor="CCFFCC")  # light green

HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT= Font(name="Calibri", bold=True, size=14, color="1E3A5F")
BOLD_FONT = Font(name="Calibri", bold=True, size=10)

thin = Side(style="thin", color="BBBBBB")
BORDER_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

def apply_header(cell, text, fill=None):
    cell.value = text
    cell.fill  = fill or HEADER_FILL
    cell.font  = HDR_FONT
    cell.alignment = CENTER
    cell.border = BORDER_THIN

def apply_input(cell, value=None, number_format=None):
    if value is not None:
        cell.value = value
    cell.fill = INPUT_FILL
    cell.font = BODY_FONT
    cell.alignment = RIGHT
    cell.border = BORDER_THIN
    if number_format:
        cell.number_format = number_format

def apply_calc(cell, formula=None, number_format=None):
    if formula is not None:
        cell.value = formula
    cell.fill = CALC_FILL
    cell.font = BODY_FONT
    cell.alignment = RIGHT
    cell.border = BORDER_THIN
    if number_format:
        cell.number_format = number_format

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Country Parameters
# ══════════════════════════════════════════════════════════════════════════════
ws_c = wb.active
ws_c.title = "Country"
ws_c.sheet_view.showGridLines = True
ws_c.column_dimensions["A"].width = 14
ws_c.column_dimensions["B"].width = 22
ws_c.column_dimensions["C"].width = 14
ws_c.column_dimensions["D"].width = 14
ws_c.column_dimensions["E"].width = 20
ws_c.column_dimensions["F"].width = 20
ws_c.column_dimensions["G"].width = 20
ws_c.column_dimensions["H"].width = 20

# Title
ws_c.merge_cells("A1:H1")
ws_c["A1"].value = "Country Parameters — Salon Financial Model"
ws_c["A1"].font  = TITLE_FONT
ws_c["A1"].alignment = CENTER
ws_c.row_dimensions[1].height = 32

# Subtitle
ws_c.merge_cells("A2:H2")
ws_c["A2"].value = "✏ Edit these cells to update assumptions. Run excel_to_json.py after any change."
ws_c["A2"].font  = Font(name="Calibri", italic=True, size=10, color="555555")
ws_c["A2"].alignment = CENTER

# Headers
headers = ["Country Code", "Country Name", "Corp. Tax Rate",
           "Currency", "Senior Stylist\n(USD/mo)",
           "Junior Stylist\n(USD/mo)", "Assistant\n(USD/mo)", "Manager\n(USD/mo)"]
for c, h in enumerate(headers, 1):
    apply_header(ws_c.cell(3, c), h)
ws_c.row_dimensions[3].height = 36

# Country data (research-backed rates in USD/month, with South Korea tax at 22.0% effective)
countries = [
    # code, name, tax, currency, senior, junior, asst, manager
    ("AU",  "Australia",   0.30, "AUD", 4225, 2800, 2470, 4550),
    ("JP",  "Japan",       0.30, "JPY", 2345, 1474, 1474, 3015),
    ("KR",  "South Korea", 0.22, "KRW", 2044, 1314, 1314, 2555),  # 20% CIT + 2% local income tax
    ("SG",  "Singapore",   0.17, "SGD", 3330, 2480, 1850, 4440),
    ("HK",  "Hong Kong",   0.165,"HKD", 2432, 1792, 1664, 3072),
    ("MC",  "Macau",       0.12, "HKD", 2304, 1536, 1536, 2816),
    ("TW",  "Taiwan",      0.20, "NTD", 1705,  992,  992, 2170),
    ("MY",  "Malaysia",    0.24, "MYR",  990,  660,  484, 1320),
    ("VN",  "Vietnam",     0.20, "VND",  800,  500,  420, 1100),
    ("TH",  "Thailand",    0.20, "THB", 1305,  870,  638, 1740),
    ("AE",  "UAE / Dubai", 0.09, "AED", 2720, 1904, 1632, 3808),
]

for r, row in enumerate(countries, 4):
    ws_c.cell(r, 1).value = row[0]
    ws_c.cell(r, 1).fill = PatternFill("solid", fgColor="E3F2FD")
    ws_c.cell(r, 1).font = BOLD_FONT
    ws_c.cell(r, 1).alignment = CENTER
    ws_c.cell(r, 1).border = BORDER_THIN
    ws_c.cell(r, 2).value = row[1]; apply_input(ws_c.cell(r, 2), row[1])
    ws_c.cell(r, 3).value = row[2]; apply_input(ws_c.cell(r, 3), row[2], "0.0%")
    ws_c.cell(r, 4).value = row[3]; apply_input(ws_c.cell(r, 4), row[3])
    for c, v in enumerate(row[4:], 5):
        apply_input(ws_c.cell(r, c), v, '"USD "[$-en-US]#,##0')

# Named range helper note
ws_c.merge_cells("A16:H16")
ws_c["A16"].value = ("💡 Country lookup: These rows are referenced by VLOOKUP in the Cities sheet. "
                      "Do not change the Country Code column (column A). Add new countries at the bottom.")
ws_c["A16"].font = Font(name="Calibri", italic=True, size=9, color="777777")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Cities (Raw Inputs)
# ══════════════════════════════════════════════════════════════════════════════
ws_i = wb.create_sheet("Cities")
ws_i.sheet_view.showGridLines = True
ws_i.freeze_panes = "C5"

# Column widths
col_widths = {
    "A": 22, "B": 14, "C": 12, "D": 12,  # identity
    "E": 10, "F": 12, "G": 12,            # space/session
    "H": 10, "I": 9,                       # ticket/cogs
    "J": 12, "K": 14, "L": 14, "M": 14, "N": 11, # rent/capex inputs
    "O": 9,  "P": 9,                       # utilisation
    "Q": 12, "R": 12, "S": 14,             # geography
    "T": 14, "U": 14                       # business/admin fees
}
for col, width in col_widths.items():
    ws_i.column_dimensions[col].width = width

# Title
ws_i.merge_cells("A1:U1")
ws_i["A1"].value = "Oasis Salon — City Input Parameters"
ws_i["A1"].font  = TITLE_FONT
ws_i["A1"].alignment = CENTER
ws_i.row_dimensions[1].height = 32

ws_i.merge_cells("A2:U2")
ws_i["A2"].value = "✏ YELLOW = editable inputs. BLUE (Model sheet) = auto-calculated. After editing, run: python excel_to_json.py"
ws_i["A2"].font  = Font(name="Calibri", italic=True, size=10, color="555555")
ws_i["A2"].alignment = CENTER
ws_i.row_dimensions[2].height = 20

# Section headers (row 3)
sections = [
    (1, 4, "IDENTITY"),
    (5, 7, "SPACE & TIME"),
    (8, 9, "PRICING"),
    (10, 14, "COSTS & CAPEX"),
    (15, 16, "UTILISATION"),
    (17, 19, "GEOGRAPHY"),
    (20, 21, "BUSINESS FEES & ADMIN")
]
for start, end, label in sections:
    ws_i.merge_cells(f"{get_column_letter(start)}3:{get_column_letter(end)}3")
    apply_header(ws_i.cell(3, start), label, SUBHDR_FILL)
ws_i.row_dimensions[3].height = 24

# Column headers (row 4)
col_headers = [
    "City Name", "URL", "Region", "Country\nCode",
    "Space\n(sq ft)", "Sessions/\nStation/Day", "Working\nDays/Month",
    "Ticket\n(USD)", "COGS %",
    "Rent\n(USD/mo)", "Fitout\n(USD)", "Equipment\n(USD)", "Other\nCAPEX (USD)", "Deposit\nMonths",
    "Util Y1\n(%)", "Util Y2\n(%)",
    "Latitude", "Longitude", "Underserved\nGap %",
    "Biz Reg Fee\n(USD)", "Admin & SW\n(USD/mo)"
]
for c, h in enumerate(col_headers, 1):
    apply_header(ws_i.cell(4, c), h)
ws_i.row_dimensions[4].height = 42

# City input data (research-backed coordinates, underserved, biz reg fees, and admin/software)
city_inputs = [
    # Name, URL, Region, Country, Sqft, Sess, Wd, Ticket, Cogs, Rent, Fitout, Equip, Other, Deposit_Mo, U1, U2, Lat, Lon, Under, Biz_Reg, Admin_SW
    ("Ho Chi Minh",       "hcmc.html",          "Vietnam",    "VN",  800, 3, 26,  95, 0.10, 3000, 20000, 20000, 9000,  3, 0.65, 0.75,  10.776,  106.701, 65, 2500, 200),
    ("Hanoi",             "hanoi.html",          "Vietnam",    "VN",  800, 3, 26,  95, 0.10, 2200, 18000, 18000, 7600,  3, 0.65, 0.75,  21.028,  105.852, 55, 2500, 200),
    ("Da Nang",           "danang.html",         "Vietnam",    "VN",  800, 4, 26,  55, 0.10,  800, 12000, 16000, 4600,  3, 0.65, 0.75,  16.068,  108.212, 70, 1500, 200),
    ("Hai Phong",         "haiphong.html",       "Vietnam",    "VN",  800, 4, 26,  45, 0.10,  650, 10000, 15000, 3300,  3, 0.65, 0.75,  20.865,  106.683, 72, 1500, 200),
    ("Binh Duong",        "binhduong.html",      "Vietnam",    "VN",  800, 4, 26,  45, 0.10,  600, 10000, 14000, 3200,  3, 0.65, 0.75,  10.980,  106.653, 75, 1500, 200),
    ("Dong Nai",          "dongnai.html",        "Vietnam",    "VN",  800, 4, 26,  55, 0.10,  550,  9000, 13000, 3100,  3, 0.65, 0.75,  10.945,  106.824, 73, 1500, 200),
    ("Kuala Lumpur",      "kuala_lumpur.html",   "Malaysia",   "MY",  800, 3, 26,  90, 0.10, 2000, 33000, 20000, 6000,  3, 0.65, 0.75,   3.148,  101.686, 40, 1500, 200),
    ("Johor Bahru (RTS)", "johor.html",          "Malaysia",   "MY", 1000, 4, 26,  65, 0.10, 1700, 28600, 18000, 5400,  3, 0.65, 0.75,   1.492,  103.743, 55, 1500, 200),
    ("Penang",            "penang.html",         "Malaysia",   "MY",  800, 4, 26,  55, 0.10,  800, 22000, 15000, 4800,  3, 0.65, 0.75,   5.414,  100.330, 50, 1000, 200),
    ("Sabah",             "sabah.html",          "Malaysia",   "MY",  800, 4, 26,  45, 0.10,  900, 19800, 14500, 4500,  3, 0.65, 0.75,   5.978,  116.073, 68, 1000, 200),
    ("Sarawak",           "sarawak.html",        "Malaysia",   "MY",  800, 4, 26,  46, 0.10,  700, 18700, 14000, 4200,  3, 0.65, 0.75,   1.553,  110.360, 65, 1000, 200),
    ("Singapore",         "singapore.html",      "Other APAC", "SG",  750, 3, 26, 180, 0.10, 5500, 66600, 22000,16500,  3, 0.65, 0.75,   1.283,  103.850, 25, 1500, 300),
    ("Hong Kong",         "hongkong.html",       "Other APAC", "HK",  750, 3, 26, 200, 0.10, 8500, 76800, 22000,25500,  3, 0.65, 0.75,  22.279,  114.165, 22, 1200, 300),
    ("Bangkok",           "bangkok.html",        "Other APAC", "TH",  800, 3, 26, 160, 0.10, 5000, 43500, 20000,15000,  3, 0.65, 0.75,  13.745,  100.532, 45, 2000, 200),
    ("Macau",             "macau.html",          "Other APAC", "MC",  750, 3, 26, 130, 0.10, 1800, 44800, 18000, 5400,  3, 0.65, 0.75,  22.197,  113.545, 38, 1500, 300),
    ("Dubai",             "dubai.html",          "Middle East", "AE", 800, 3, 26, 160, 0.10, 3500, 54400, 22000,42000, 12, 0.65, 0.75,  25.187,   55.274, 30, 6000, 300),
    ("Taipei",            "taipei.html",         "Taiwan",     "TW",  850, 3, 26, 104, 0.10, 2200, 55800, 20000, 6600,  3, 0.65, 0.75,  25.033,  121.565, 35, 1800, 200),
    ("Taichung",          "taichung.html",       "Taiwan",     "TW",  800, 3, 26, 110, 0.10, 1800, 46500, 18000, 5400,  3, 0.65, 0.75,  24.163,  120.647, 42, 1800, 200),
    ("Kaohsiung",         "kaohsiung.html",      "Taiwan",     "TW",  800, 3, 26, 120, 0.10, 1200, 37200, 17000, 3600,  3, 0.65, 0.75,  22.627,  120.302, 48, 1800, 200),
    ("Tainan",            "tainan.html",         "Taiwan",     "TW",  750, 3, 26, 100, 0.10,  900, 34100, 16000, 2700,  3, 0.65, 0.75,  22.993,  120.202, 55, 1800, 200),
    ("Fukuoka",           "fukuoka.html",        "Japan",      "JP",  800, 3, 22, 130, 0.10, 1600,100500, 20000, 4800,  3, 0.65, 0.75,  33.589,  130.398, 38, 2500, 300),
    ("Okinawa",           "okinawa.html",        "Japan",      "JP",  800, 3, 22, 120, 0.10, 1100, 80400, 18000, 3300,  3, 0.65, 0.75,  26.213,  127.679, 52, 2500, 300),  # Japan shops closed Mon = 22 days/mo
    ("Busan",             "busan.html",          "South Korea","KR",  800, 3, 26, 100, 0.10, 2200, 58400, 18000, 6600,  3, 0.65, 0.75,  35.158,  129.060, 45, 2000, 300),
    ("Sydney",            "sydney.html",         "Australia",  "AU",  800, 3, 26, 250, 0.10, 3200, 91000, 22000, 9600,  3, 0.65, 0.75, -33.869,  151.207, 30, 1500, 300),
    ("Melbourne",         "melbourne.html",      "Australia",  "AU",  800, 3, 26, 220, 0.10, 3500, 87750, 22000,10500,  3, 0.65, 0.75, -37.814,  144.963, 32, 1500, 300),
    ("Brisbane",          "brisbane.html",       "Australia",  "AU",  800, 3, 26, 240, 0.10, 2800, 78000, 20000, 8400,  3, 0.65, 0.75, -27.470,  153.026, 36, 1500, 300),
    ("Perth",             "perth.html",          "Australia",  "AU",  800, 3, 26, 200, 0.10, 2000, 74750, 19000, 6000,  3, 0.65, 0.75, -31.953,  115.857, 38, 1500, 300),
]

for r_idx, row in enumerate(city_inputs, 5):
    ws_i.cell(r_idx, 1).value = row[0];    ws_i.cell(r_idx, 1).fill = PatternFill("solid", fgColor="E3F2FD"); ws_i.cell(r_idx, 1).font = BOLD_FONT; ws_i.cell(r_idx, 1).alignment = LEFT;  ws_i.cell(r_idx, 1).border = BORDER_THIN
    ws_i.cell(r_idx, 2).value = row[1];    apply_input(ws_i.cell(r_idx, 2), row[1]); ws_i.cell(r_idx, 2).alignment = LEFT
    ws_i.cell(r_idx, 3).value = row[2];    apply_input(ws_i.cell(r_idx, 3), row[2]); ws_i.cell(r_idx, 3).alignment = LEFT
    ws_i.cell(r_idx, 4).value = row[3];    apply_input(ws_i.cell(r_idx, 4), row[3]); ws_i.cell(r_idx, 4).alignment = CENTER
    apply_input(ws_i.cell(r_idx,  5), row[4],   "#,##0")   # Space
    apply_input(ws_i.cell(r_idx,  6), row[5],   "0")       # Sessions
    apply_input(ws_i.cell(r_idx,  7), row[6],   "0")       # Working Days
    apply_input(ws_i.cell(r_idx,  8), row[7],   '"$"#,##0') # Ticket
    apply_input(ws_i.cell(r_idx,  9), row[8],   "0.0%")    # COGS%
    apply_input(ws_i.cell(r_idx, 10), row[9],   '"$"#,##0') # Rent
    apply_input(ws_i.cell(r_idx, 11), row[10],  '"$"#,##0') # Fitout
    apply_input(ws_i.cell(r_idx, 12), row[11],  '"$"#,##0') # Equip
    apply_input(ws_i.cell(r_idx, 13), row[12],  '"$"#,##0') # Other
    apply_input(ws_i.cell(r_idx, 14), row[13],  "0")       # Deposit Months
    apply_input(ws_i.cell(r_idx, 15), row[14],  "0%")      # Util Y1
    apply_input(ws_i.cell(r_idx, 16), row[15],  "0%")      # Util Y2
    apply_input(ws_i.cell(r_idx, 17), row[16],  "0.000")   # Lat
    apply_input(ws_i.cell(r_idx, 18), row[17],  "0.000")   # Lon
    apply_input(ws_i.cell(r_idx, 19), row[18],  "0")       # Underserved
    apply_input(ws_i.cell(r_idx, 20), row[19],  '"$"#,##0') # Biz Reg Fee
    apply_input(ws_i.cell(r_idx, 21), row[20],  '"$"#,##0') # Admin & SW OPEX


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Model (All calculations)
# ══════════════════════════════════════════════════════════════════════════════
ws_m = wb.create_sheet("Model")
ws_m.sheet_view.showGridLines = True
ws_m.freeze_panes = "C5"

# Column widths for 38 columns (A to AL)
model_widths = {
    "A": 22,  # City name
    "B": 10,  # Country
    "C": 8,   "D": 10,  "E": 8,             # Stations, Color Bar Seats, Total Seats
    "F": 8,   "G": 8,   "H": 8,   "I": 10,  "J": 10, # Staff
    "K": 10,  "L": 14,  "M": 14,  "N": 14,  "O": 14, # Country Rates
    "P": 14,  # Total Staff Cost
    "Q": 12,  "R": 12,  "S": 12,  "T": 14,  "U": 14, "V": 16, # OPEX (Rent, Utilities, Marketing, Admin/SW, Deprec, Cash OPEX)
    "W": 12,  "X": 11,  "Y": 11,            # Capacity (Max, Y1, Y2)
    "Z": 14,  "AA": 13, "AB": 13,           # Economics (BE, COGS/sess, GM/sess)
    "AC": 14, "AD": 14, "AE": 14, "AF": 14, # Profit & Cash Flow (Y1 PAT, Y2 PAT, Y1 CF, Y2 CF)
    "AG": 12, "AH": 12, "AI": 12, "AJ": 13, # CAPEX & Payback (Low, High, Mid, Payback months)
    "AK": 12, "AL": 12                      # Ratio, Viable
}
for col, width in model_widths.items():
    ws_m.column_dimensions[col].width = width

# Title
ws_m.merge_cells("A1:AL1")
ws_m["A1"].value = "Oasis Salon — Calculated Financial Model (Auto-Generated from Cities Sheet)"
ws_m["A1"].font  = TITLE_FONT
ws_m["A1"].alignment = CENTER
ws_m.row_dimensions[1].height = 32

ws_m.merge_cells("A2:AL2")
ws_m["A2"].value = "🔒 Do not edit this sheet directly — change inputs in Cities sheet. BLUE = formula cells."
ws_m["A2"].font  = Font(name="Calibri", italic=True, size=10, color="CC0000")
ws_m["A2"].alignment = CENTER

# Section headers (row 3)
model_sections = [
    (1, 2, "IDENTITY"),
    (3, 5, "SPACE & SEATS"),
    (6, 10, "STAFFING MODEL"),
    (11, 15, "COUNTRY RATES"),
    (16, 16, "STAFF COST"),
    (17, 22, "MONTHLY OPEX"),
    (23, 25, "CAPACITY"),
    (26, 28, "UNIT ECONOMICS"),
    (29, 32, "PROFIT & CASH FLOW"),
    (33, 36, "CAPEX & PAYBACK"),
    (37, 37, "RATIO"),
    (38, 38, "VIABLE?"),
]
for start, end, label in model_sections:
    ws_m.merge_cells(f"{get_column_letter(start)}3:{get_column_letter(end)}3")
    apply_header(ws_m.cell(3, start), label, SUBHDR_FILL)
ws_m.row_dimensions[3].height = 24

# Column headers (row 4)
model_headers = [
    "City Name", "Country",
    "Stations", "Color Bar\nSeats", "Total\nSeats",
    "Wash\nBasins", "Seniors", "Juniors", "Assistants", "Total\nStaff",
    "Tax Rate", "Senior\n(USD/mo)", "Junior\n(USD/mo)", "Assistant\n(USD/mo)", "Manager\n(USD/mo)",
    "Total Staff\nCost (USD/mo)",
    "Rent\n(USD/mo)", "Utilities\n(USD/mo)", "Mktg+Misc\n(USD/mo)", "Admin & SW\n(USD/mo)", "Monthly\nDeprec (USD)", "Total Cash\nOPEX (USD/mo)",
    "Max\nCapacity", "Y1 Clients\n(65%)", "Y2 Clients\n(75%)",
    "Breakeven\n(clients/mo)", "COGS/session", "GM/session\n(90%)",
    "Y1 Monthly\nPAT (USD)", "Y2 Monthly\nPAT (USD)",
    "Y1 Monthly\nCashFlow (USD)", "Y2 Monthly\nCashFlow (USD)",
    "CAPEX Low", "CAPEX High", "CAPEX Mid", "Payback\n(months)",
    "PAT/OPEX\n(%)",
    "VIABLE?\n(BE < 80%cap)",
]
for c, h in enumerate(model_headers, 1):
    apply_header(ws_m.cell(4, c), h)
ws_m.row_dimensions[4].height = 48

# Formula rows (rows 5 to 31)
for r_idx in range(5, 32):
    ci = r_idx  # Cities row index
    cr = r_idx  # Model row index
    
    def c_col(col_letter):
        return f"Cities!{col_letter}{ci}"
        
    def vlookup(col_num):
        return f"VLOOKUP({c_col('D')},Country!$A$4:$H$14,{col_num},FALSE)"

    # Identity
    ws_m.cell(cr, 1).value  = f"=Cities!A{ci}"
    ws_m.cell(cr, 1).fill   = PatternFill("solid", fgColor="E3F2FD")
    ws_m.cell(cr, 1).font   = BOLD_FONT
    ws_m.cell(cr, 1).alignment = LEFT
    ws_m.cell(cr, 1).border = BORDER_THIN
    
    ws_m.cell(cr, 2).value = f"={c_col('D')}"
    ws_m.cell(cr, 2).fill  = PatternFill("solid", fgColor="E3F2FD")
    ws_m.cell(cr, 2).font  = BOLD_FONT
    ws_m.cell(cr, 2).alignment = CENTER
    ws_m.cell(cr, 2).border = BORDER_THIN

    # Space & Seats
    apply_calc(ws_m.cell(cr, 3), f"=IF({c_col('E')}<=750,3,IF({c_col('E')}<=850,4,5))", "0") # Stations
    apply_calc(ws_m.cell(cr, 4), f"=C{cr}-1", "0") # Color Bar Seats
    apply_calc(ws_m.cell(cr, 5), f"=C{cr}+D{cr}", "0") # Total Seats

    # Staffing Model
    apply_calc(ws_m.cell(cr, 6), f"=IF(C{cr}<=4,2,3)", "0") # Wash Basins
    apply_calc(ws_m.cell(cr, 7), f"=IF(C{cr}<=4,2,3)", "0") # Seniors
    apply_calc(ws_m.cell(cr, 8), f"=C{cr}-G{cr}", "0") # Juniors
    apply_calc(ws_m.cell(cr, 9), f"=CEILING(C{cr}/2,1)", "0") # Assistants
    apply_calc(ws_m.cell(cr, 10), f"=G{cr}+H{cr}+I{cr}+1", "0") # Total Staff

    # Country Rates
    apply_calc(ws_m.cell(cr, 11), f"={vlookup(3)}", "0.0%") # Tax
    apply_calc(ws_m.cell(cr, 12), f"={vlookup(5)}", '"$"#,##0') # Senior rate
    apply_calc(ws_m.cell(cr, 13), f"={vlookup(6)}", '"$"#,##0') # Junior rate
    apply_calc(ws_m.cell(cr, 14), f"={vlookup(7)}", '"$"#,##0') # Assistant rate
    apply_calc(ws_m.cell(cr, 15), f"={vlookup(8)}", '"$"#,##0') # Manager rate

    # Total Staff Cost
    apply_calc(ws_m.cell(cr, 16), f"=G{cr}*L{cr}+H{cr}*M{cr}+I{cr}*N{cr}+O{cr}", '"$"#,##0')

    # Monthly OPEX
    apply_calc(ws_m.cell(cr, 17), f"={c_col('J')}", '"$"#,##0') # Rent
    apply_calc(ws_m.cell(cr, 18), f"=200+C{cr}*80", '"$"#,##0') # Utilities
    apply_calc(ws_m.cell(cr, 19), f"=350+C{cr}*50+300+C{cr}*30", '"$"#,##0') # Marketing/Misc
    apply_calc(ws_m.cell(cr, 20), f"={c_col('U')}", '"$"#,##0') # Admin & SW
    apply_calc(ws_m.cell(cr, 21), f"={c_col('L')}/60", '"$"#,##0') # Monthly Depreciation (Equipment/60)
    apply_calc(ws_m.cell(cr, 22), f"=Q{cr}+P{cr}+R{cr}+S{cr}+T{cr}", '"$"#,##0') # Total Cash OPEX

    # Capacity
    apply_calc(ws_m.cell(cr, 23), f"=C{cr}*{c_col('F')}*{c_col('G')}", "0") # Max Cap
    apply_calc(ws_m.cell(cr, 24), f"=INT(W{cr}*{c_col('O')})", "0") # Y1 Clients
    apply_calc(ws_m.cell(cr, 25), f"=INT(W{cr}*{c_col('P')})", "0") # Y2 Clients

    # Unit Economics
    apply_calc(ws_m.cell(cr, 26), f"=CEILING((V{cr}+U{cr})/AB{cr},1)", "0") # Breakeven Accounting
    apply_calc(ws_m.cell(cr, 27), f"={c_col('H')}*{c_col('I')}", '"$"#,##0.00') # COGS
    apply_calc(ws_m.cell(cr, 28), f"={c_col('H')}-AA{cr}", '"$"#,##0.00') # GM

    # Profitability & Cash Flow
    # Y1 Monthly PAT: (Y1_clients * GM - Cash_OPEX - Depreciation) * (1 - Tax)
    apply_calc(ws_m.cell(cr, 29),
               f"=IF((X{cr}*AB{cr}-V{cr}-U{cr})>0, (X{cr}*AB{cr}-V{cr}-U{cr})*(1-K{cr}), X{cr}*AB{cr}-V{cr}-U{cr})", '"$"#,##0')
    # Y2 Monthly PAT
    apply_calc(ws_m.cell(cr, 30),
               f"=IF((Y{cr}*AB{cr}-V{cr}-U{cr})>0, (Y{cr}*AB{cr}-V{cr}-U{cr})*(1-K{cr}), Y{cr}*AB{cr}-V{cr}-U{cr})", '"$"#,##0')
    # Y1 Monthly Cash Flow = Y1_PAT + Depreciation
    apply_calc(ws_m.cell(cr, 31), f"=AC{cr}+U{cr}", '"$"#,##0')
    # Y2 Monthly Cash Flow = Y2_PAT + Depreciation
    apply_calc(ws_m.cell(cr, 32), f"=AD{cr}+U{cr}", '"$"#,##0')

    # CAPEX & Payback
    # CAPEX Low: (fitout+equip+other+biz_reg+rent*deposit) * 0.92
    apply_calc(ws_m.cell(cr, 33),
               f"=FLOOR(({c_col('K')}+{c_col('L')}+{c_col('M')}+{c_col('T')}+{c_col('J')}*{c_col('N')})*0.92,5000)", '"$"#,##0')
    # CAPEX High
    apply_calc(ws_m.cell(cr, 34),
               f"=CEILING(({c_col('K')}+{c_col('L')}+{c_col('M')}+{c_col('T')}+{c_col('J')}*{c_col('N')})*1.08,5000)", '"$"#,##0')
    # CAPEX Mid
    apply_calc(ws_m.cell(cr, 35), f"=(AG{cr}+AH{cr})/2", '"$"#,##0')
    # Payback using cumulative cash flow (using Y1 Cash Flow for first 12 mo, then Y2)
    apply_calc(ws_m.cell(cr, 36),
               f"=IF(AE{cr}*12>=AI{cr}, IF(AE{cr}>0,CEILING(AI{cr}/AE{cr},1),999), IF(AF{cr}>0,12+CEILING((AI{cr}-AE{cr}*12)/AF{cr},1),999))", "0")

    # Ratio & Viability
    apply_calc(ws_m.cell(cr, 37), f"=IF(V{cr}>0,AD{cr}/V{cr},0)", "0%") # PAT / OPEX Ratio
    
    # Viable? Accounting breakeven < 80% capacity
    cell_vi = ws_m.cell(cr, 38)
    cell_vi.value = f"=IF(Z{cr}<W{cr}*0.8,\"✓ YES\",\"⚠ TIGHT\")"
    cell_vi.fill  = CALC_FILL
    cell_vi.font  = BOLD_FONT
    cell_vi.alignment = CENTER
    cell_vi.border = BORDER_THIN

# Conditional formatting for Viable column
from openpyxl.formatting.rule import CellIsRule
ws_m.conditional_formatting.add(
    f"AL5:AL31",
    CellIsRule(operator="containsText", formula=['"✓"'], fill=OK_FILL, font=Font(color="006100", bold=True))
)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Instructions
# ══════════════════════════════════════════════════════════════════════════════
ws_inst = wb.create_sheet("Instructions")
ws_inst.column_dimensions["A"].width = 110

instructions = [
    ("Oasis Salon Feasibility — Excel Workbook Guide", "title"),
    ("", ""),
    ("📋 HOW THIS WORKBOOK IS STRUCTURED", "section"),
    ("", ""),
    ("• Country sheet:    Edit tax rates and monthly staff salaries (in USD) per country.", "body"),
    ("• Cities sheet:     Edit all raw inputs per city (yellow cells only):", "body"),
    ("   - Space (sq ft), sessions/day, working days, ticket price, COGS %", "sub"),
    ("   - Rent, fitout cost, equipment cost, other CAPEX, deposit months", "sub"),
    ("   - Latitude, Longitude, and Underserved Gap % (for map visualization)", "sub"),
    ("   - Biz Reg Fee (one-time license fee) and Admin & SW (monthly admin overhead)", "sub"),
    ("   - Year 1 and Year 2 utilisation targets", "sub"),
    ("• Model sheet:      All calculations. Do NOT edit directly. Reads from Cities + Country.", "body"),
    ("", ""),
    ("⚙️ KEY FORMULAS", "section"),
    ("", ""),
    ("  Styling Stations = IF(sqft ≤ 750 → 3, sqft ≤ 850 → 4, else 5)", "body"),
    ("  Color Bar Seats  = Stations - 1  (Provides overflow capacity for concurrent color processing)", "body"),
    ("  Total Seats      = Stations + Color Bar Seats", "body"),
    ("  Max Capacity     = Stations × sessions/day × working_days/month", "body"),
    ("  Total Staff Cost = seniors×senior_rate + juniors×junior_rate + assistants×asst_rate + manager_rate", "body"),
    ("  Depreciation     = Equipment Cost / 60  (Straight-line over 5 years / 60 months)", "body"),
    ("  Total Cash OPEX  = Rent + Staff_Cost + Utilities(200+stations×80) + Mktg+Misc(650+stations×80) + Admin/SW", "body"),
    ("  Accounting BE    = CEILING((Total Cash OPEX + Depreciation) / (ticket × (1 − COGS%)), 1)", "body"),
    ("  CAPEX Base       = Fitout + Equipment + Other + Biz_Reg_Fee + Rent × Deposit_Months", "body"),
    ("  Y2 Monthly PAT   = (Y2_clients × ticket × (1−COGS%) − Cash OPEX − Depreciation) × (1 − tax_rate)", "body"),
    ("  Y2 Monthly CF    = Y2_PAT + Depreciation (Add back non-cash depreciation)", "body"),
    ("  Payback          = If Y1 Cash Flow covers CAPEX: CAPEX/Y1_CF. Else: 12 + (CAPEX - Y1_CF*12)/Y2_CF", "body"),
    ("  Viable?          = YES if Accounting Breakeven < 80% of Max Capacity", "body"),
    ("", ""),
    ("🔄 UPDATE WORKFLOW", "section"),
    ("", ""),
    ("  1. Edit values in Country or Cities sheet (yellow cells).", "body"),
    ("  2. Review the Model sheet to check that all formulas recalculate correctly.", "body"),
    ("  3. Save the Excel file.", "body"),
    ("  4. Open a terminal and run:", "body"),
    ("       python excel_to_json.py", "code"),
    ("  5. This regenerates city_data.json.", "body"),
    ("  6. Commit and push:", "body"),
    ("       git add -A && git commit -m \"Update financials\" && git push", "code"),
    ("", ""),
    ("📌 NOTES", "section"),
    ("", ""),
    ("  • Taipei vs Taichung CAPEX: Taipei (USD 80-100k) > Taichung (USD 65-85k) because", "body"),
    ("    Taipei has: higher rent ($2,200 vs $1,800) + higher Da'an district fitout (USD 55.8k vs 46.5k) + 850 sqft vs 800 sqft.", "sub"),
    ("  • Japan (Fukuoka, Okinawa) uses 22 working days/month due to standard Monday closures in the Japanese market.", "body"),
    ("  • Dubai uses 12 deposit months (UAE annual lease pre-payment norm).", "body"),
    ("  • South Korea corporate tax is set to 22.0% (incorporating 20.0% national + 2.0% local income taxes).", "body"),
    ("  • Biz Reg Fee covers company setup, legal counsel, bank creation, and local salon permit fees.", "body"),
    ("  • Equipment depreciation over 5 years reduces taxable income (tax shield) and is added back for cash flow payback.", "body"),
]

style_map = {
    "title":   (TITLE_FONT, 20),
    "section": (Font(name="Calibri", bold=True, size=12, color="1E3A5F"), 18),
    "body":    (Font(name="Calibri", size=10), 15),
    "sub":     (Font(name="Calibri", size=10, color="555555"), 14),
    "code":    (Font(name="Courier New", size=10, color="006400"), 14),
    "":        (Font(name="Calibri", size=10), 8),
}

for r_idx, (text, style) in enumerate(instructions, 1):
    cell = ws_inst.cell(r_idx, 1)
    cell.value = text
    font, height = style_map.get(style, (BODY_FONT, 14))
    cell.font = font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_inst.row_dimensions[r_idx].height = height

# ── ORDER SHEETS ──────────────────────────────────────────────────────────────
wb.move_sheet("Instructions", offset=-10)

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUT_FILE)
print(f"✅ Created: {OUT_FILE}")
print(f"   Sheets: {[ws.title for ws in wb.worksheets]}")
print(f"   Cities: {len(city_inputs)} rows")
print(f"   Countries: {len(countries)} rows")
