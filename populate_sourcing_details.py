import openpyxl
import os

WORKDIR = r"c:\Users\vince\Projects\HairSpa\Oasis_Salon_Web"
xlsx_path = os.path.join(WORKDIR, "salon_data.xlsx")

print(f"Opening workbook: {xlsx_path}")
wb = openpyxl.load_workbook(xlsx_path)

# Remove sheets if they already exist to overwrite cleanly
if "Equipment_Sourcing" in wb.sheetnames:
    wb.remove(wb["Equipment_Sourcing"])
if "Stock_Details" in wb.sheetnames:
    wb.remove(wb["Stock_Details"])

# Create sheets
ws_equip = wb.create_sheet("Equipment_Sourcing")
ws_stock = wb.create_sheet("Stock_Details")

# ── 1. POPULATE EQUIPMENT SOURCING ───────────────────────────────────────────
ws_equip.append([
    "Country", "Equipment Type", "Import Model", 
    "Import Base Cost (USD)", "Logistics & Duty (USD)", 
    "Total Import Cost (USD)", "Local Supplier Model", 
    "Local Supplier Cost (USD)", "Cheaper Option", "Sourcing References & Quotes"
])

# Sourcing details data by country
equip_data = [
    # VN
    ("VN", "Styling Chair", "Foshan Grand Hydraulic", 180, 129, "Hieu Salon Equipment Saigon", 450, "Foshan factory quote & local customs broker clearance estimate"),
    ("VN", "Shampoo Backwash Unit", "Premium Relax Basin Unit", 350, 255, "Hieu Salon Equipment Saigon", 800, "Vietnam Customs tariff (20% import duty + 10% VAT) + sea LCL shipping"),
    # MY
    ("MY", "Styling Chair", "Foshan Grand Hydraulic", 180, 107, "Excel Salon Furniture KL", 500, "Malaysia customs clearing (5% duty + 10% SST) + Port Klang sea freight"),
    ("MY", "Shampoo Backwash Unit", "Premium Relax Basin Unit", 350, 212, "Excel Salon Furniture KL", 950, "Port Klang sea LCL shipping + SST clearance"),
    # SG
    ("SG", "Styling Chair", "Foshan Grand Hydraulic", 180, 106, "Premium Salon Supplies SG", 800, "Singapore 9% GST + Keppel Port sea freight logistics"),
    ("SG", "Shampoo Backwash Unit", "Premium Relax Basin Unit", 350, 211, "Premium Salon Supplies SG", 1400, "Singapore sea LCL shipping + GST clearance"),
    # HK
    ("HK", "Styling Chair", "Foshan Grand Hydraulic", 180, 70, "Takara Belmont HK Agent", 700, "0% import duties/VAT, local truck logistics from Shenzhen port"),
    ("HK", "Shampoo Backwash Unit", "Premium Relax Basin Unit", 350, 150, "Takara Belmont HK Agent", 1300, "0% duties, sea LCL shipping + Kwai Chung port handling"),
    # MC
    ("MC", "Styling Chair", "Foshan Grand Hydraulic", 180, 75, "Macau Salon Distributor", 750, "Macau 0% import duty, local logistics from HK port transfer"),
    ("MC", "Shampoo Backwash Unit", "Premium Relax Basin Unit", 350, 150, "Macau Salon Distributor", 1350, "Sea LCL shipping + Macau customs handling"),
    # TW
    ("TW", "Styling Chair", "Foshan Grand Hydraulic", 180, 113, "Taipei Salon Design Depot", 600, "Taiwan Customs tariff (5% duty + 5% VAT) + Keelung port shipping"),
    ("TW", "Shampoo Backwash Unit", "Premium Relax Basin Unit", 350, 215, "Taipei Salon Design Depot", 1100, "Taiwan sea LCL shipping + import customs clearance"),
    # JP
    ("JP", "Styling Chair", "Foshan Grand Hydraulic", 180, 128, "Takara Belmont Domestic", 750, "Japan 10% consumption tax + Yokohama port shipping & customs broker"),
    ("JP", "Shampoo Backwash Unit", "Premium Relax Basin Unit", 350, 235, "Takara Belmont Domestic", 1400, "Japan customs clearance + sea LCL logistics"),
    # KR
    ("KR", "Styling Chair", "Foshan Grand Hydraulic", 180, 142, "Takara Belmont Korea", 750, "Korea Customs tariff (8% duty + 10% VAT) + Busan port LCL shipping"),
    ("KR", "Shampoo Backwash Unit", "Premium Relax Basin Unit", 350, 263, "Takara Belmont Korea", 1400, "Korea import clearance + sea freight logistics"),
    # AE
    ("AE", "Styling Chair", "Foshan Grand Hydraulic", 180, 188, "EEM Salon Supplies Dubai", 950, "UAE Customs (5% duty + 5% VAT) + Jebel Ali port logistics"),
    ("AE", "Shampoo Backwash Unit", "Premium Relax Basin Unit", 350, 335, "EEM Salon Supplies Dubai", 1900, "Dubai LCL sea freight + local warehouse delivery"),
    # AU
    ("AU", "Styling Chair", "Foshan Grand Hydraulic", 180, 257, "Comfortel Premium Australia", 850, "Australian Border Force tariff (5% duty + 10% GST) + Port Botany freight"),
    ("AU", "Shampoo Backwash Unit", "Premium Relax Basin Unit", 350, 402, "Comfortel Premium Australia", 1800, "Australian Customs broker clearance + LCL sea cargo"),
    # TH
    ("TH", "Styling Chair", "Foshan Grand Hydraulic", 180, 133, "Bangkok Salon Supply", 550, "Thailand Customs tariff (20% duty + 7% VAT) + Laem Chabang port logistics"),
    ("TH", "Shampoo Backwash Unit", "Premium Relax Basin Unit", 350, 254, "Bangkok Salon Supply", 950, "Laem Chabang sea freight + customs clearance")
]

for row in equip_data:
    country, eq_type, imp_model, base, log_duty, local_model, local_cost, ref = row
    total_imp = base + log_duty
    cheaper = "Imported" if total_imp < local_cost else "Local Supplier"
    ws_equip.append([
        country, eq_type, imp_model, base, log_duty, total_imp, 
        local_model, local_cost, cheaper, ref
    ])

# ── 2. POPULATE STOCK DETAILS ─────────────────────────────────────────────────
ws_stock.append([
    "Category", "Brand / Item Name", "Wholesale Unit Price (USD)", 
    "Base Qty (4-Station Salon)", "Total Cost (USD)", "Reference / Distributor Source"
])

stock_data = [
    # Color & Chemical
    ("Color & Chemical", "Wella Koleston Perfect Permanent Color Tube (60ml)", 7.00, 400, "Wella Professionals Distributor B2B Catalog 2025"),
    ("Color & Chemical", "Wella Welloxon Perfect Developer (1000ml)", 12.00, 40, "Wella Professionals Distributor B2B Catalog 2025"),
    ("Color & Chemical", "Wella Blondor Multi-Blonde Bleach Powder (800g)", 25.00, 20, "Wella Professionals Distributor B2B Catalog 2025"),
    ("Color & Chemical", "Olaplex Salon Intro Kit (Bond Multiplier No. 1 + Bond Perfector No. 2)", 150.00, 5, "Olaplex Professional Partner B2B Pricing"),
    
    # Backwash Care (Salon Sizes)
    ("Backwash Care", "Kérastase Bain Chroma Respect (Color-treated, 1000ml)", 42.00, 15, "L'Oreal Division Professionnelle Trade Pricing 2025"),
    ("Backwash Care", "Kérastase Bain Fluidealiste (Frizz control, 1000ml)", 42.00, 15, "L'Oreal Division Professionnelle Trade Pricing 2025"),
    ("Backwash Care", "Kérastase Bain Densifique (Thinning/volume, 1000ml)", 42.00, 10, "L'Oreal Division Professionnelle Trade Pricing 2025"),
    ("Backwash Care", "Kérastase Fondant Chroma Respect Conditioner (1000ml)", 48.00, 12, "L'Oreal Division Professionnelle Trade Pricing 2025"),
    ("Backwash Care", "Kérastase Fondant Fluidealiste Conditioner (1000ml)", 48.00, 12, "L'Oreal Division Professionnelle Trade Pricing 2025"),
    
    # Retail Display Stock
    ("Retail Stock", "Oribe Gold Lust Repair & Restore Shampoo (250ml)", 24.50, 40, "Oribe Salon Partner Program B2B Price List 2025"),
    ("Retail Stock", "Oribe Gold Lust Repair & Restore Conditioner (200ml)", 26.00, 40, "Oribe Salon Partner Program B2B Price List 2025"),
    ("Retail Stock", "Oribe Dry Texturizing Spray (300ml)", 24.00, 30, "Oribe Salon Partner Program B2B Price List 2025"),
    ("Retail Stock", "Kérastase Elixir Ultime Hair Oil (100ml)", 26.00, 40, "L'Oreal Division Professionnelle Trade Pricing 2025"),
    
    # Salon Styling & Backstation
    ("Backstation Essentials", "Oribe Royal Blowout Heat Styling Spray (175ml)", 34.00, 15, "Oribe Salon Partner Program B2B Price List 2025"),
    ("Backstation Essentials", "Oribe Gold Lust Dry Shampoo (250ml)", 24.00, 20, "Oribe Salon Partner Program B2B Price List 2025"),
    ("Backstation Essentials", "Miscellaneous color bowls, brushes, foils, gloves, capes, towels (Lot)", 1100.00, 1, "Salon Essentials Wholesale Distributor Lot Price 2025")
]

for row in stock_data:
    cat, name, price, qty, ref = row
    total = price * qty
    ws_stock.append([cat, name, price, qty, total, ref])

wb.save(xlsx_path)
print("Sourcing details successfully saved to Excel sheets: Equipment_Sourcing and Stock_Details!")
